from __future__ import annotations

import logging
import os
import json
import tempfile
import csv
from io import BytesIO, StringIO
from pathlib import Path

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Response, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from app.auth import RequestContext, get_request_context
from app.config import app_env, is_production
from app.db import build_repository
from app.models import AccessRequest, CloseoutEvidence, Comment, Item, ItemCreate, ItemStatus, ItemUpdate, ProjectConfig, RectificationEvidence, RAISED_BY_OPTIONS, Settings, SubProfile, TRADES
from app.permissions import (
    require_close_item,
    require_comment_access,
    require_create_item,
    require_demo_reset,
    require_issue_item,
    require_item_access,
    require_rectification_access,
    require_report_access,
    require_storage_status_access,
    require_update_item,
    visible_items,
    visible_projects,
)
from app.services import items as item_service
from app.services import projects as project_service
from app.services import reports as report_service
from app.storage import StorageUploadError, resolve_photo_url
from app.validation import ValidationError
from app.workflow import WorkflowError

logger = logging.getLogger(__name__)
APP_DIR = Path(__file__).resolve().parent
REPO_ROOT = APP_DIR.parent
STATIC_DIR = APP_DIR / "static"
LEGACY_EXPORT_DIR = REPO_ROOT / "CleanRun-IQ-Full-App-Render3"
LEGACY_ASSETS_DIR = LEGACY_EXPORT_DIR / "assets"


app = FastAPI(title="CleanRun IQ Python", version="0.1.0")
store = build_repository()

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
if LEGACY_ASSETS_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(LEGACY_ASSETS_DIR)), name="assets")


class IssuePayload(BaseModel):
    to: str
    by: str | None = None
    note: str | None = None
    reissue: bool = False


class ActorPayload(BaseModel):
    by: str = "Site Team"
    note: str | None = None


class RejectPayload(BaseModel):
    by: str = "Site Team"
    reason: str


class RectificationPayload(BaseModel):
    photo: str | None = None
    comment: str | None = None
    by: str | None = None
    advance_to_ready: bool = False


class SettingsPayload(BaseModel):
    active_project: str | None = None
    company: str | None = None
    prepared_by: str | None = None
    projects: list[str] | None = None
    project_configs: dict[str, ProjectConfig] | None = None
    subcontractors: list[str] | None = None
    sub_profiles: dict[str, SubProfile] | None = None


class LegacyParsePayload(BaseModel):
    transcript: str | None = None
    text: str | None = None


def _normalise_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


async def spreadsheet_rows(file: UploadFile) -> list[list[str]]:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="Spreadsheet is empty")
    filename = (file.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(status_code=503, detail="Excel import dependency is unavailable") from exc
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [[str(cell or "").strip() for cell in row] for row in sheet.iter_rows(values_only=True)]
        return [row for row in rows if any(row)]
    text = raw.decode("utf-8-sig", errors="replace")
    delimiter = "\t" if filename.endswith(".tsv") else ","
    rows = [[cell.strip() for cell in row] for row in csv.reader(StringIO(text), delimiter=delimiter)]
    return [row for row in rows if any(row)]


def import_units_from_rows(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    headers = [_normalise_header(cell) for cell in rows[0]]
    unit_indexes = [idx for idx, header in enumerate(headers) if header in {"unit", "units", "area", "areas", "unitarea", "unitareas", "apartment", "apartments", "lot", "lots"}]
    data_rows = rows[1:] if unit_indexes else rows
    indexes = unit_indexes or list(range(max(len(row) for row in rows)))
    values: list[str] = []
    for row in data_rows:
        for idx in indexes:
            if idx < len(row) and row[idx].strip():
                values.append(row[idx].strip())
    return sorted(set(values), key=lambda value: value.lower())


def import_subcontractors_from_rows(rows: list[list[str]]) -> tuple[list[str], dict[str, SubProfile]]:
    if not rows:
        return [], {}
    headers = [_normalise_header(cell) for cell in rows[0]]
    name_headers = {"name", "subcontractor", "subcontractors", "company", "business", "contractor"}
    header_map = {header: idx for idx, header in enumerate(headers)}
    name_idx = next((idx for idx, header in enumerate(headers) if header in name_headers), None)
    has_headers = name_idx is not None
    data_rows = rows[1:] if has_headers else rows
    profiles: dict[str, SubProfile] = {}

    def cell(row: list[str], *names: str) -> str | None:
        for name in names:
            idx = header_map.get(name)
            if idx is not None and idx < len(row) and row[idx].strip():
                return row[idx].strip()
        return None

    for row in data_rows:
        name = row[name_idx].strip() if has_headers and name_idx is not None and name_idx < len(row) else (row[0].strip() if row else "")
        if not name:
            continue
        profiles[name] = SubProfile(
            name=name,
            trade=cell(row, "trade", "discipline", "category"),
            contact=cell(row, "contact", "contactname", "representative"),
            email=cell(row, "email", "emailaddress"),
            phone=cell(row, "phone", "mobile", "telephone"),
        )
    return sorted(profiles, key=lambda value: value.lower()), profiles


def actor_label(ctx: RequestContext) -> str:
    return ctx.user.audit_label


def actor_context(ctx: RequestContext) -> dict[str, str | None]:
    return {
        "id": ctx.user.id,
        "email": ctx.user.email,
        "role": ctx.user.company_role or ctx.user.project_roles.get("*"),
        "auth_method": ctx.user.auth_method,
    }


def get_authorized_item(item_id: str, ctx: RequestContext):
    try:
        item = item_service.get_item(store, item_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    require_item_access(ctx.user, item)
    return item


def scoped_settings(settings: Settings, ctx: RequestContext) -> Settings:
    projects = visible_projects(ctx.user, settings)
    project_configs = {name: cfg for name, cfg in settings.project_configs.items() if name in projects}
    active_project = settings.active_project if settings.active_project in projects else (projects[0] if projects else "")
    return settings.model_copy(
        update={
            "projects": projects,
            "project_configs": project_configs,
            "active_project": active_project,
            "prepared_by": ctx.user.audit_label,
        }
    )


def camel_settings(settings: Settings) -> dict[str, object]:
    payload = settings.model_dump(mode="json")
    payload["activeProject"] = payload.pop("active_project", "")
    payload["projectConfigs"] = payload.pop("project_configs", {})
    payload["subProfiles"] = payload.pop("sub_profiles", {})
    payload["preparedBy"] = payload.pop("prepared_by", "")
    for config in payload["projectConfigs"].values():
        if "default_due_days" in config:
            config["defaultDueDays"] = config.pop("default_due_days")
        if "preferred_items_view" in config:
            config["preferredItemsView"] = config.pop("preferred_items_view")
        if "code_prefix" in config:
            config["codePrefix"] = config.pop("code_prefix")
        if "code_prefix_locked" in config:
            config["codePrefixLocked"] = config.pop("code_prefix_locked")
        if "code_prefix_hidden_on_cards" in config:
            config["codePrefixHiddenOnCards"] = config.pop("code_prefix_hidden_on_cards")
    return payload


def sign_item_photos(item: Item) -> Item:
    original_photos = [resolve_photo_url(photo) or photo for photo in item.original_photos]
    rectification_evidence = [
        evidence.model_copy(
            update={"photo": (resolve_photo_url(evidence.photo) or evidence.photo) if evidence.photo else evidence.photo}
        )
        for evidence in item.rectification_evidence
    ]
    closeout_evidence = [
        evidence.model_copy(
            update={"photo": (resolve_photo_url(evidence.photo) or evidence.photo) if evidence.photo else evidence.photo}
        )
        for evidence in item.closeout_evidence
    ]
    return item.model_copy(
        update={
            "original_photos": original_photos,
            "rectification_evidence": rectification_evidence,
            "closeout_evidence": closeout_evidence,
        }
    )


def camel_item(item) -> dict[str, object]:
    payload = sign_item_photos(item).model_dump(mode="json")
    rename = {
        "due_date": "dueDate",
        "original_photos": "originalPhotos",
        "voice_transcript": "voiceTranscript",
        "voice_note": "voiceNote",
        "created_by": "createdBy",
        "created_at": "createdAt",
        "updated_at": "updatedAt",
        "rectification_evidence": "rectificationEvidence",
        "closeout_evidence": "closeoutEvidence",
        "issue_history": "issueHistory",
        "inspection_history": "inspectionHistory",
        "audit_events": "auditEvents",
        "issued_at": "issuedAt",
        "in_progress_at": "inProgressAt",
        "ready_for_review_at": "readyForReviewAt",
        "under_inspection_at": "underInspectionAt",
        "closed_at": "closedAt",
        "rejection_reason": "rejectionReason",
    }
    for source, target in rename.items():
        if source in payload:
            payload[target] = payload.pop(source)
    return payload


def visible_project_items(ctx: RequestContext, items: list[Item], project_name: str) -> list[Item]:
    project_items = [item for item in items if item.project == project_name]
    return visible_items(ctx.user, project_items)


def user_payload(ctx: RequestContext, *, camel: bool = False) -> dict[str, object]:
    if camel:
        return {
            "id": ctx.user.id,
            "email": ctx.user.email,
            "companyRole": ctx.user.company_role,
            "projectRoles": ctx.user.project_roles,
            "subcontractors": sorted(ctx.user.subcontractors),
        }
    return {
        "id": ctx.user.id,
        "email": ctx.user.email,
        "company_role": ctx.user.company_role,
        "project_roles": ctx.user.project_roles,
        "subcontractors": sorted(ctx.user.subcontractors),
    }


def workflow_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, WorkflowError):
        return HTTPException(status_code=409, detail=str(exc))
    if isinstance(exc, ValidationError):
        return HTTPException(status_code=422, detail=str(exc))
    if isinstance(exc, StorageUploadError):
        return HTTPException(status_code=413, detail=str(exc))
    raise exc


def snake_item_payload(payload: dict[str, object]) -> dict[str, object]:
    rename = {
        "dueDate": "due_date",
        "originalPhotos": "original_photos",
        "voiceTranscript": "voice_transcript",
        "voiceNote": "voice_note",
        "createdBy": "created_by",
        "raisedBy": "raised_by",
        "appendOriginalPhotos": "append_original_photos",
    }
    result = dict(payload)
    for source, target in rename.items():
        if source in result:
            result[target] = result.pop(source)
    return result


def snake_settings_payload(payload: dict[str, object]) -> dict[str, object]:
    result = dict(payload)
    if "activeProject" in result:
        result["active_project"] = result.pop("activeProject")
    if "projectConfigs" in result:
        result["project_configs"] = result.pop("projectConfigs")
    if "subProfiles" in result:
        result["sub_profiles"] = result.pop("subProfiles")
    if "preparedBy" in result:
        result["prepared_by"] = result.pop("preparedBy")
    configs = result.get("project_configs")
    if isinstance(configs, dict):
        for config in configs.values():
            if isinstance(config, dict):
                if "defaultDueDays" in config:
                    config["default_due_days"] = config.pop("defaultDueDays")
                if "preferredItemsView" in config:
                    config["preferred_items_view"] = config.pop("preferredItemsView")
                if "codePrefix" in config:
                    config["code_prefix"] = config.pop("codePrefix")
                if "codePrefixLocked" in config:
                    config["code_prefix_locked"] = config.pop("codePrefixLocked")
                if "codePrefixHiddenOnCards" in config:
                    config["code_prefix_hidden_on_cards"] = config.pop("codePrefixHiddenOnCards")
    return result


def _match_config_value(text: str, values: list[str]) -> str | None:
    lowered = text.lower()
    for value in values:
        if value and value.lower() in lowered:
            return value
    return None


def configured_voice_context(project: str, ctx: RequestContext) -> dict[str, object]:
    require_create_item(ctx.user, project)
    data = store.snapshot()
    visible = scoped_settings(data.settings, ctx)
    cfg = visible.project_configs.get(project)
    if not cfg:
        raise HTTPException(status_code=400, detail="Unknown or unauthorized project for voice parsing")
    return {
        "project": project,
        "buildings": cfg.buildings,
        "levels": cfg.levels,
        "units": cfg.units,
        "rooms": cfg.rooms,
        "trades": TRADES,
        "subcontractors": visible.subcontractors,
        "sub_profiles": {name: profile.model_dump() for name, profile in visible.sub_profiles.items()},
        "item_types": ["defect", "incomplete", "client"],
        "priorities": ["high", "urgent"],
    }


def voice_item_schema() -> dict[str, object]:
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "project": {"type": ["string", "null"]},
            "building": {"type": ["string", "null"]},
            "level": {"type": ["string", "null"]},
            "unit": {"type": ["string", "null"]},
            "room": {"type": ["string", "null"]},
            "trade": {"type": ["string", "null"]},
            "subcontractor": {"type": ["string", "null"]},
            "priority": {"type": "string", "enum": ["high", "urgent"]},
            "type": {"type": "string", "enum": ["defect", "incomplete", "client"]},
            "due_date": {"type": ["string", "null"]},
            "description": {"type": "string"},
            "raw_transcript": {"type": "string"},
            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
            "warnings": {"type": "array", "items": {"type": "string"}},
        },
        "required": [
            "project",
            "building",
            "level",
            "unit",
            "room",
            "trade",
            "subcontractor",
            "priority",
            "type",
            "due_date",
            "description",
            "raw_transcript",
            "confidence",
            "warnings",
        ],
    }


@app.get("/api/auth/config")
def auth_config() -> dict[str, object]:
    return {
        "supabase_url": os.getenv("SUPABASE_URL"),
        "supabase_publishable_key": os.getenv("SUPABASE_PUBLISHABLE_KEY"),
        "environment": os.getenv("CLEANRUN_ENV", "development"),
        "dev_tokens_enabled": not is_production(),
    }


@app.post("/api/access-requests", status_code=201)
def create_access_request(payload: AccessRequest):
    try:
        return store.create_access_request(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception as exc:
        logger.exception("Access request submission failed for %s", payload.email)
        raise HTTPException(
            status_code=503,
            detail="Could not submit request. Please email info@cleanruniq.com.",
        ) from exc


@app.get("/api/deploy")
def deploy_status() -> dict[str, object]:
    return {
        "app": "cleanrun-iq",
        "version": app.version,
        "render_git_commit": os.getenv("RENDER_GIT_COMMIT"),
        "render_service_name": os.getenv("RENDER_SERVICE_NAME"),
        "storage": os.getenv("CLEANRUN_STORAGE", "local"),
        "supabase_url_configured": bool(os.getenv("SUPABASE_URL")),
    }


@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    legacy_app = LEGACY_EXPORT_DIR / "index.html"
    if legacy_app.exists():
        html = legacy_app.read_text(encoding="utf-8")
        return HTMLResponse(html)
    html = (STATIC_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(html)


@app.get("/manifest.webmanifest")
def manifest() -> Response:
    manifest_path = LEGACY_EXPORT_DIR / "manifest.webmanifest"
    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail="Manifest not found")
    return Response(manifest_path.read_text(encoding="utf-8"), media_type="application/manifest+json")


@app.get("/service-worker.js")
def service_worker() -> Response:
    worker_path = LEGACY_EXPORT_DIR / "service-worker.js"
    if not worker_path.exists():
        raise HTTPException(status_code=404, detail="Service worker not found")
    return Response(worker_path.read_text(encoding="utf-8"), media_type="application/javascript")


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/health")
def api_health() -> dict[str, bool]:
    return {"ok": True}


@app.get("/api/voice/status")
def voice_status(ctx: RequestContext = Depends(get_request_context)) -> dict[str, bool | str]:
    return {
        "ai_voice_enabled": bool(os.getenv("OPENAI_API_KEY")),
        "transcribe_model": os.getenv("OPENAI_TRANSCRIBE_MODEL", "gpt-4o-mini-transcribe"),
        "parse_model": os.getenv("OPENAI_PARSE_MODEL", "gpt-4o-mini"),
    }


@app.post("/api/voice/parse")
async def parse_voice_note(
    audio: UploadFile = File(...),
    project: str = Form(...),
    ctx: RequestContext = Depends(get_request_context),
):
    if not os.getenv("OPENAI_API_KEY"):
        raise HTTPException(status_code=503, detail="AI voice parsing is not configured. Type the note and use Draft form from note.")

    try:
        from openai import OpenAI
    except Exception as exc:
        logger.exception("OpenAI SDK unavailable")
        raise HTTPException(status_code=503, detail="AI voice parser dependency is unavailable.") from exc

    context = configured_voice_context(project, ctx)
    suffix = Path(audio.filename or "voice-note.webm").suffix or ".webm"
    audio_bytes = await audio.read()
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="Voice recording was empty. Retry or type the note.")

    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(audio_bytes)
            tmp_path = tmp.name

        with open(tmp_path, "rb") as audio_file:
            transcription = client.audio.transcriptions.create(
                model=os.getenv("OPENAI_TRANSCRIBE_MODEL", "gpt-4o-mini-transcribe"),
                file=audio_file,
                response_format="text",
            )

        transcript = transcription if isinstance(transcription, str) else getattr(transcription, "text", "")
        transcript = str(transcript or "").strip()
        if not transcript:
            raise HTTPException(status_code=422, detail="No speech was detected. Retry or type the note.")

        completion = client.chat.completions.create(
            model=os.getenv("OPENAI_PARSE_MODEL", "gpt-4o-mini"),
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You convert Australian construction site voice notes into CleanRun IQ item fields. "
                        "Use the allowed project context wherever possible. Do not invent locations, trades, or subcontractors. "
                        "Preserve the full spoken note as raw_transcript, but keep description concise and remove structured location data. "
                        "Example: 'Building 3, Unit 305, Balcony, on Level 1, tiling to be repaired.' becomes "
                        "building B3/Building 3, unit U305/Unit 305, room Balcony, level Level 1, trade Tiling, "
                        "description 'Tiling to be repaired.'"
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps({"transcript": transcript, "allowed_context": context}),
                },
            ],
            response_format={
                "type": "json_schema",
                "json_schema": {
                    "name": "cleanrun_voice_item",
                    "strict": True,
                    "schema": voice_item_schema(),
                },
            },
        )
        parsed = json.loads(completion.choices[0].message.content or "{}")
        parsed["raw_transcript"] = transcript
        return {"transcript": transcript, "parsed": parsed, "source": "ai"}
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("AI voice parse failed")
        raise HTTPException(status_code=502, detail="AI voice parsing failed. Retry or type the note.") from exc
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@app.post("/api/parse")
def legacy_parse_note(payload: LegacyParsePayload, ctx: RequestContext = Depends(get_request_context)):
    text = (payload.transcript or payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=422, detail="Type or speak a note first.")

    data = store.snapshot()
    settings = scoped_settings(data.settings, ctx)
    project = settings.active_project or (settings.projects[0] if settings.projects else "")
    cfg = settings.project_configs.get(project)
    parsed: dict[str, object] = {"description": text, "project": project}
    if cfg:
        parsed.update(
            {
                "building": _match_config_value(text, cfg.buildings),
                "level": _match_config_value(text, cfg.levels),
                "unit": _match_config_value(text, cfg.units),
                "room": _match_config_value(text, cfg.rooms),
            }
        )
    parsed["trade"] = _match_config_value(text, TRADES)
    return {key: value for key, value in parsed.items() if value}


@app.get("/api/storage-status")
def storage_status(ctx: RequestContext = Depends(get_request_context)):
    require_storage_status_access(ctx.user)
    data = store.snapshot()
    response = {
        "requested_storage": os.getenv("CLEANRUN_STORAGE", "local"),
        "active_store": store.__class__.__name__,
        "environment": app_env(),
        "supabase_url_configured": bool(os.getenv("SUPABASE_URL")),
        "supabase_publishable_key_configured": bool(os.getenv("SUPABASE_PUBLISHABLE_KEY")),
        "auth_jwt_secret_configured": bool(os.getenv("SUPABASE_JWT_SECRET")),
        "requires_supabase": os.getenv("CLEANRUN_STORAGE", "").lower() == "supabase",
        "storage_bucket": os.getenv("CLEANRUN_STORAGE_BUCKET", "cleanrun-evidence"),
        "item_count": len(data.items),
    }
    if not is_production():
        latest = data.items[0] if data.items else None
        response["latest_item_code"] = latest.code if latest else None
    return response


@app.get("/api/bootstrap")
def bootstrap(ctx: RequestContext = Depends(get_request_context)):
    data = store.snapshot()
    items = visible_items(ctx.user, data.items)
    return {
        "settings": scoped_settings(data.settings, ctx),
        "items": [sign_item_photos(item) for item in items],
        "trades": TRADES,
        "raised_by_options": RAISED_BY_OPTIONS,
        "user": user_payload(ctx),
    }


@app.get("/api/state")
def legacy_state(ctx: RequestContext = Depends(get_request_context)):
    data = store.snapshot()
    settings = scoped_settings(data.settings, ctx)
    return {
        "settings": camel_settings(settings),
        "items": [camel_item(item) for item in visible_items(ctx.user, data.items)],
        "plans": [],
        "trades": TRADES,
        "raisedByOptions": RAISED_BY_OPTIONS,
        "user": user_payload(ctx, camel=True),
    }


@app.patch("/api/settings")
def update_settings(payload: SettingsPayload, ctx: RequestContext = Depends(get_request_context)):
    require_storage_status_access(ctx.user)
    data = store.snapshot()
    current = data.settings
    updates = payload.model_dump(exclude_unset=True)
    settings = Settings.model_validate({**current.model_dump(), **updates})
    try:
        return project_service.update_settings(store, settings)
    except project_service.SettingsLockError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Settings save failed for user=%s", ctx.user.email)
        raise HTTPException(status_code=503, detail="Could not save settings. Check Render logs for the Supabase write error.") from exc


@app.post("/api/settings")
def legacy_update_settings(payload: dict[str, object], ctx: RequestContext = Depends(get_request_context)):
    return update_settings(SettingsPayload.model_validate(snake_settings_payload(payload)), ctx)


@app.post("/api/settings/import")
async def import_settings_spreadsheet(
    target: str = Form(...),
    project: str | None = Form(default=None),
    file: UploadFile = File(...),
    ctx: RequestContext = Depends(get_request_context),
):
    require_storage_status_access(ctx.user)
    data = store.snapshot()
    settings = data.settings
    rows = await spreadsheet_rows(file)
    target_key = target.strip().lower()
    updates: dict[str, object]
    imported = 0

    if target_key in {"units", "unit", "areas", "unit_areas"}:
        project_name = project or settings.active_project
        cfg = settings.project_configs.get(project_name)
        if not cfg:
            raise HTTPException(status_code=404, detail="Project not found")
        values = import_units_from_rows(rows)
        merged = sorted(set([*cfg.units, *values]), key=lambda value: value.lower())
        project_configs = dict(settings.project_configs)
        project_configs[project_name] = cfg.model_copy(update={"units": merged})
        updates = {"project_configs": project_configs}
        imported = len([value for value in values if value not in cfg.units])
    elif target_key in {"subcontractors", "subcontractor", "subs"}:
        names, profiles = import_subcontractors_from_rows(rows)
        existing_names = set(settings.subcontractors)
        sub_profiles = dict(settings.sub_profiles)
        for name, profile in profiles.items():
            current = sub_profiles.get(name)
            sub_profiles[name] = profile if not current else current.model_copy(
                update={key: value for key, value in profile.model_dump().items() if value}
            )
        subcontractors = sorted(set([*settings.subcontractors, *names]), key=lambda value: value.lower())
        updates = {"subcontractors": subcontractors, "sub_profiles": sub_profiles}
        imported = len([name for name in names if name not in existing_names])
    else:
        raise HTTPException(status_code=422, detail="Import target must be units or subcontractors")

    next_settings = Settings.model_validate({**settings.model_dump(), **updates})
    saved = project_service.update_settings(store, next_settings)
    return {"settings": saved, "imported": imported, "target": target_key}


@app.get("/api/items")
def list_items(
    project: str | None = Query(default=None),
    status: str | None = Query(default=None),
    ctx: RequestContext = Depends(get_request_context),
):
    if project:
        require_report_access(ctx.user, project)
    return visible_items(ctx.user, item_service.list_items(store, project=project, status=status))


@app.get("/api/items/{item_id}")
def get_item(item_id: str, ctx: RequestContext = Depends(get_request_context)):
    return get_authorized_item(item_id, ctx)


@app.post("/api/items", status_code=201)
def create_item(payload: dict[str, object], issue_now: bool = Query(default=False), ctx: RequestContext = Depends(get_request_context)):
    payload = ItemCreate.model_validate(snake_item_payload(payload))
    require_create_item(ctx.user, payload.project)
    payload = payload.model_copy(update={"created_by": actor_label(ctx)})
    try:
        return item_service.create_item(store, payload, issue_now=issue_now, actor=actor_context(ctx))
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except StorageUploadError as exc:
        raise HTTPException(status_code=413, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Item create failed for user=%s project=%s issue_now=%s", ctx.user.email, payload.project, issue_now)
        raise HTTPException(status_code=503, detail="Could not save item. Check Render logs for the Supabase write error.") from exc


@app.patch("/api/items/{item_id}")
def update_item(item_id: str, payload: dict[str, object], by: str | None = Query(default=None), ctx: RequestContext = Depends(get_request_context)):
    payload = ItemUpdate.model_validate(snake_item_payload(payload))
    item = get_authorized_item(item_id, ctx)
    require_update_item(ctx.user, item)
    if payload.project is not None and payload.project != item.project:
        require_create_item(ctx.user, payload.project)
    try:
        return item_service.update_item(store, item_id, payload, by=actor_label(ctx), actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/actions/{action}")
def legacy_item_action(item_id: str, action: str, payload: dict[str, object], ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    if action == "in-progress":
        action = "start"
    try:
        if action == "issue":
            require_issue_item(ctx.user, item)
            return camel_item(store.issue_item(item_id, to=str(payload.get("to") or item.subcontractor), by=actor_label(ctx), note=payload.get("note"), reissue=bool(payload.get("reissue")), actor=actor_context(ctx)))
        if action == "start":
            require_rectification_access(ctx.user, item)
            return camel_item(store.mark_in_progress(item_id, by=actor_label(ctx), actor=actor_context(ctx)))
        if action == "ready":
            require_rectification_access(ctx.user, item)
            return camel_item(store.mark_ready(item_id, by=actor_label(ctx), actor=actor_context(ctx)))
        if action == "inspect":
            require_close_item(ctx.user, item)
            return camel_item(store.start_inspection(item_id, by=actor_label(ctx), actor=actor_context(ctx)))
        if action == "reject":
            require_close_item(ctx.user, item)
            return camel_item(store.reject(item_id, by=actor_label(ctx), reason=str(payload.get("reason") or ""), actor=actor_context(ctx)))
        if action == "close":
            require_close_item(ctx.user, item)
            evidence = CloseoutEvidence(
                photo=payload.get("photo"),
                by=actor_label(ctx),
                role=str(payload.get("role") or "Supervisor"),
                note=payload.get("note"),
                confirmation=str(payload.get("confirmation") or payload.get("confirmed") or ""),
            )
            return camel_item(store.close_with_evidence(item_id, evidence, actor=actor_context(ctx)))
        if action == "rectification":
            require_rectification_access(ctx.user, item)
            evidence = RectificationEvidence(photo=payload.get("photo"), comment=payload.get("comment"), by=actor_label(ctx))
            return camel_item(store.add_rectification(item_id, evidence, advance_to_ready=bool(payload.get("advanceToReady") or payload.get("advance_to_ready")), actor=actor_context(ctx)))
        if action == "comment":
            require_comment_access(ctx.user, item)
            comment = Comment(text=str(payload.get("text") or ""), by=actor_label(ctx))
            return camel_item(store.add_comment(item_id, comment, actor=actor_context(ctx)))
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)
    except StorageUploadError as exc:
        raise HTTPException(status_code=413, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Item action failed for user=%s item=%s action=%s", ctx.user.email, item_id, action)
        raise HTTPException(status_code=503, detail="Could not update item. Check Render logs for the Supabase write error.") from exc
    raise HTTPException(status_code=404, detail="Unknown item action")


@app.post("/api/items/{item_id}/issue")
def issue_item(item_id: str, payload: IssuePayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_issue_item(ctx.user, item)
    try:
        return store.issue_item(item_id, to=payload.to, by=actor_label(ctx), note=payload.note, reissue=payload.reissue, actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/in-progress")
def mark_in_progress(item_id: str, payload: ActorPayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_rectification_access(ctx.user, item)
    try:
        return store.mark_in_progress(item_id, by=actor_label(ctx), actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/ready")
def mark_ready(item_id: str, payload: ActorPayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_rectification_access(ctx.user, item)
    try:
        return store.mark_ready(item_id, by=actor_label(ctx), actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/inspection/start")
def start_inspection(item_id: str, payload: ActorPayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_close_item(ctx.user, item)
    try:
        return store.start_inspection(item_id, by=actor_label(ctx), actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/inspection/reject")
def reject_item(item_id: str, payload: RejectPayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_close_item(ctx.user, item)
    try:
        return store.reject(item_id, by=actor_label(ctx), reason=payload.reason, actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/closeout")
def closeout_item(item_id: str, payload: CloseoutEvidence, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_close_item(ctx.user, item)
    payload = payload.model_copy(update={"by": actor_label(ctx)})
    try:
        return store.close_with_evidence(item_id, payload, actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/rectification")
def add_rectification(item_id: str, payload: RectificationPayload, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_rectification_access(ctx.user, item)
    try:
        evidence = RectificationEvidence(photo=payload.photo, comment=payload.comment, by=actor_label(ctx))
        return store.add_rectification(item_id, evidence, advance_to_ready=payload.advance_to_ready, actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")
    except (WorkflowError, ValidationError) as exc:
        raise workflow_http_error(exc)


@app.post("/api/items/{item_id}/comments")
def add_comment(item_id: str, payload: Comment, ctx: RequestContext = Depends(get_request_context)):
    item = get_authorized_item(item_id, ctx)
    require_comment_access(ctx.user, item)
    payload = payload.model_copy(update={"by": actor_label(ctx)})
    try:
        return store.add_comment(item_id, payload, actor=actor_context(ctx))
    except KeyError:
        raise HTTPException(status_code=404, detail="Item not found")


@app.get("/api/reports/{report_type}", response_class=HTMLResponse)
def report_html(
    report_type: str,
    project: str | None = Query(default=None),
    subcontractor: str | None = Query(default=None),
    ctx: RequestContext = Depends(get_request_context),
):
    data = store.snapshot()
    project_name = project or data.settings.active_project
    require_report_access(ctx.user, project_name)
    items = visible_project_items(ctx, data.items, project_name)
    settings = data.settings.model_copy(update={"active_project": project_name})
    html = report_service.build_report(items, settings, report_type=report_type, subcontractor=subcontractor)
    return HTMLResponse(html)


@app.get("/api/reports/{report_type}/summary")
def report_summary(
    report_type: str,
    project: str | None = Query(default=None),
    subcontractor: str | None = Query(default=None),
    ctx: RequestContext = Depends(get_request_context),
):
    data = store.snapshot()
    project_name = project or data.settings.active_project
    require_report_access(ctx.user, project_name)
    items = report_service.report_items(visible_project_items(ctx, data.items, project_name), report_type, subcontractor=subcontractor)
    return {
        "report_type": report_type,
        "project": project_name,
        "count": len(items),
        "closed": len([i for i in items if i.status in {ItemStatus.CLOSED, ItemStatus.COMPLETE}]),
        "outstanding": len([i for i in items if i.status not in {ItemStatus.CLOSED, ItemStatus.COMPLETE}]),
    }


@app.post("/api/reset-demo")
def reset_demo(ctx: RequestContext = Depends(get_request_context)):
    require_demo_reset(ctx.user)
    return store.reset_demo()


@app.post("/api/reset")
def legacy_reset_demo(ctx: RequestContext = Depends(get_request_context)):
    return reset_demo(ctx)
